from utils.Configuration import Configuration
import requests
from openpyxl import Workbook
import json


def LerBase():
    file = open("./db/municipios.json","r", encoding='utf-8-sig')
    municipios = json.load(file)
    return municipios

def ObterMunicipio(municipios,estados,wb):
    for municipio in municipios:
        uf = municipio["codigo_uf"]
        estado = ObterEstado(uf,estados)
        objMunicipio = type('',(object,),{"nome": municipio["nome"],"estado":estado})()
        EscreverMunicipio(objMunicipio,wb)

def ObterEstado(uf,estados):
    for estado in estados:
       if estado["id"] == uf:
        return estado["nome"]

def ObterEstados(config):
    request = requests.get(config.base_url+"localidades/estados")
    estados = request.json()
    return estados

def CriarArquivo(estados):
    wb = Workbook()
    #wb = wb.active
    for estado in estados:
        nome = estado["nome"]
        ws = wb.create_sheet(nome)
        ws.title = nome
    wb.save('./prefeituras.xlsx')
    return wb


def EscreverMunicipio(municipio,wb):
    querystring =f'https://www.google.com/search?q=contatos+prefeitura+{municipio.nome}+estado+de+{municipio.estado}'
    sheet = wb[municipio.estado]
    row = sheet.max_row+1
    sheet.cell(column=1, row=row, value=municipio.nome)
    sheet.cell(column=2, row=row, value=municipio.estado)
    sheet.cell(column=3, row=row, value='=HYPERLINK("{}", "{}")'.format(querystring, f'Link {municipio.nome}'))


if __name__ == "__main__":
    config = Configuration()
    estados = ObterEstados(config)
    wb = CriarArquivo(estados)
    municipios = LerBase()
    ObterMunicipio(municipios,estados,wb)
    wb.save('./prefeituras.xlsx')
